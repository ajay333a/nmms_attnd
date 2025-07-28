import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import time
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import io
from attendance_downloader import get_attendance_data, download_photo
from openpyxl.drawing.image import Image as XLImage
import re
from concurrent.futures import ThreadPoolExecutor

# Constants
BASE_URL = "https://mnregaweb4.nic.in/nregaarch/View_NMMS_atten_date_new.aspx?fin_year=2024-2025&Digest=HNrisV4bhHnb7Gve3mAKYQ"
STATE_VALUE = '15'  # Karnataka
DISTRICT_NAME = 'BALLARI'
BLOCK_NAME = 'SIRUGUPPA'
TALUK_NAME = 'Siruguppa'
DISTRICT_LABEL = 'Ballari'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

# Fills for conditional formatting
PRESENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
ABSENT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
MALE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")    # Blue
FEMALE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Yellow

def get_table_by_id_or_div(soup, table_id='grdTable', div_id='RepPr1'):
    table = soup.find('table', {'id': table_id})
    if not table:
        div = soup.find('div', {'id': div_id})
        if div:
            table = div.find('table')
    return table

def get_link_from_table(table, match_col_idx, match_text, href_col_idx=None):
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) > match_col_idx and cols[match_col_idx].get_text(strip=True).upper() == match_text.upper():
            a = cols[match_col_idx].find('a', href=True)
            if a:
                return a['href']
    return None

def get_panchayath_link(table, panchayath_name):
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) >= 4:
            s_no = cols[0].get_text(strip=True)
            if not s_no.isdigit():
                continue
            panch_name = cols[1].get_text(strip=True).upper()
            muster_rolls_a = cols[3].find('a', href=True)
            href = muster_rolls_a['href'] if muster_rolls_a else None
            if panch_name == panchayath_name and href:
                return href
    return None

def get_muster_roll_rows(muster_table, choice, workcodes=None, workcode_idx=None, muster_no_idx=None):
    rows_to_save = []
    for row in muster_table.find_all('tr')[1:]:
        cols = row.find_all('td')
        if len(cols) > muster_no_idx:
            if choice == 'all':
                muster_a = cols[muster_no_idx].find('a', href=True)
                if muster_a:
                    muster_href = muster_a['href']
                    rows_to_save.append((cols, muster_href))
            elif choice == 'work' and workcodes and workcode_idx is not None:
                if len(cols) > workcode_idx:
                    cell_text = cols[workcode_idx].get_text(strip=True)
                    match = re.search(r'(\d+/[A-Z]+/\S+)', cell_text)
                    if match:
                        work_code_in_row = match.group(1).strip()
                        if work_code_in_row in workcodes:
                            muster_a = cols[muster_no_idx].find('a', href=True)
                            if muster_a:
                                muster_href = muster_a['href']
                                rows_to_save.append((cols, muster_href))
    return rows_to_save

def prompt_for_work_selection(work_codes, panchayath_name):
    print("\nAvailable work codes for", panchayath_name + ":")
    for code in work_codes:
        print(code)
    choice = input("\nType 'all' for all muster rolls or 'work' for specific work: ").strip().lower()
    workcodes = []
    if choice == 'work':
        workcode_input = input("Enter one or more workcodes, separated by commas: ").strip()
        workcodes = [code.strip() for code in workcode_input.split(',')]
    return choice, workcodes

def save_attendance_excel(wb, img_wb, rfo_wb):
    """Saves workbooks to memory and returns them as BytesIO objects."""
    wb_bytes = io.BytesIO()
    wb.save(wb_bytes)
    wb_bytes.seek(0)

    img_wb_bytes = io.BytesIO()
    img_wb.save(img_wb_bytes)
    img_wb_bytes.seek(0)

    rfo_wb_bytes = io.BytesIO()
    rfo_wb.save(rfo_wb_bytes)
    rfo_wb_bytes.seek(0)
    
    return wb_bytes, img_wb_bytes, rfo_wb_bytes

def save_raw_excel(rows_to_save, panchayath_name, attendance_date, muster_no_idx, workcode_idx, panchayath_url, muster_data_cache):
    raw_wb = openpyxl.Workbook()
    raw_ws = raw_wb.active
    raw_ws.append([
        "Taluk", "Panchayath", "Work Code", "Muster Roll No", "Job Card No", "Worker Name", "Gender", "Attendance", "Attendance Date", "Photo Taken By"
    ])
    for cols, muster_href in rows_to_save:
        muster_url = urljoin(panchayath_url, muster_href)
        attendance_data, _, _, _, taken_by = muster_data_cache.get(muster_url, (None, None, None, None, None))
        muster_roll_no = cols[muster_no_idx].get_text(strip=True)
        work_code = cols[workcode_idx].get_text(strip=True)
        for att_row in attendance_data or []:
            worker_name_full = att_row[2] if len(att_row) > 2 else ''
            if worker_name_full.endswith(')') and '(' in worker_name_full:
                name_part = worker_name_full[:worker_name_full.rfind('(')].strip()
                gender_part = worker_name_full[worker_name_full.rfind('(')+1:-1].strip()
            else:
                name_part = worker_name_full
                gender_part = ''
            
            attendance_status = att_row[4] if len(att_row) > 4 else ''
            
            raw_ws.append([
                TALUK_NAME,
                panchayath_name,
                work_code,
                muster_roll_no,
                att_row[1] if len(att_row) > 1 else '',
                name_part,
                gender_part,
                attendance_status,
                att_row[3] if len(att_row) > 3 else '',
                taken_by
            ])

            # Apply conditional formatting to the row just added
            current_row = raw_ws.max_row
            gender_cell = raw_ws[f'G{current_row}']
            attendance_cell = raw_ws[f'H{current_row}']

            if gender_part.upper() == 'F':
                gender_cell.fill = FEMALE_FILL
            elif gender_part.upper() == 'M':
                gender_cell.fill = MALE_FILL
            
            if attendance_status.upper() == 'P':
                attendance_cell.fill = PRESENT_FILL
            elif attendance_status.upper() == 'A':
                attendance_cell.fill = ABSENT_FILL

    raw_wb_bytes = io.BytesIO()
    raw_wb.save(raw_wb_bytes)
    raw_wb_bytes.seek(0)
    return raw_wb_bytes

def find_col_idx(header_cols, search):
    search_clean = re.sub(r'[^a-zA-Z0-9]', '', search.lower())
    for i, h in enumerate(header_cols):
        h_clean = re.sub(r'[^a-zA-Z0-9]', '', h.lower())
        if search_clean in h_clean:
            return i
    return None

def fetch_muster_data(muster_url):
    return get_attendance_data(muster_url)

def get_available_dates():
    """Fetches the list of available attendance dates from the website."""
    try:
        session = requests.Session()
        resp = session.get(BASE_URL, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')
        attendance_select = soup.find('select', {'name': 'ctl00$ContentPlaceHolder1$ddl_attendance'})
        if not attendance_select:
            return []
        
        # The first option is usually '--Select--', so we skip it.
        date_options = [opt['value'] for opt in attendance_select.find_all('option') if opt.get('value')]
        return date_options
    except requests.RequestException as e:
        print(f"Error fetching dates: {e}")
        return []

def get_panchayath_and_work_codes(attendance_date, panchayath_name):
    """Navigates to the panchayath page and extracts available work codes."""
    panchayath_name = panchayath_name.upper()
    try:
        session = requests.Session()
        resp = session.get(BASE_URL, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(resp.content, 'html.parser')

        viewstate_tag = soup.find('input', {'id': '__VIEWSTATE'})
        eventvalidation_tag = soup.find('input', {'id': '__EVENTVALIDATION'})
        viewstategen_tag = soup.find('input', {'id': '__VIEWSTATEGENERATOR'})

        if not all([viewstate_tag, eventvalidation_tag, viewstategen_tag]):
            print("Could not find required form elements on the page.")
            return None

        data = {
            '__VIEWSTATE': viewstate_tag['value'],
            '__VIEWSTATEGENERATOR': viewstategen_tag['value'],
            '__EVENTVALIDATION': eventvalidation_tag['value'],
            'ctl00$ContentPlaceHolder1$ddlstate': STATE_VALUE,
            'ctl00$ContentPlaceHolder1$ddl_attendance': attendance_date,
            'ctl00$ContentPlaceHolder1$btn_showreport': 'Show Attendance',
        }
        headers_post = HEADERS.copy()
        headers_post['Referer'] = BASE_URL
        resp2 = session.post(BASE_URL, data=data, headers=headers_post, timeout=30)
        soup2 = BeautifulSoup(resp2.content, 'html.parser')

        state_table = get_table_by_id_or_div(soup2)
        if not state_table: return None
        karnataka_link = get_link_from_table(state_table, 1, 'KARNATAKA')
        if not karnataka_link: return None
        karnataka_url = urljoin(BASE_URL, karnataka_link)

        resp3 = session.get(karnataka_url, headers=HEADERS, timeout=15)
        soup3 = BeautifulSoup(resp3.content, 'html.parser')
        dist_table = get_table_by_id_or_div(soup3)
        if not dist_table: return None
        ballari_link = get_link_from_table(dist_table, 1, DISTRICT_NAME)
        if not ballari_link: return None
        ballari_url = urljoin(karnataka_url, ballari_link)

        resp4 = session.get(ballari_url, headers=HEADERS, timeout=15)
        soup4 = BeautifulSoup(resp4.content, 'html.parser')
        block_table = get_table_by_id_or_div(soup4)
        if not block_table: return None
        siruguppa_link = get_link_from_table(block_table, 1, BLOCK_NAME)
        if not siruguppa_link: return None
        siruguppa_url = urljoin(ballari_url, siruguppa_link)

        resp5 = session.get(siruguppa_url, headers=HEADERS, timeout=15)
        soup5 = BeautifulSoup(resp5.content, 'html.parser')
        panch_div = soup5.find('div', {'id': 'RepPr1'})
        if not panch_div: return None
        panch_table = panch_div.find('table')
        if not panch_table: return None
        panchayath_link = get_panchayath_link(panch_table, panchayath_name)
        if not panchayath_link: return None
        panchayath_url = urljoin(siruguppa_url, panchayath_link)

        resp6 = session.get(panchayath_url, headers=HEADERS, timeout=15)
        soup6 = BeautifulSoup(resp6.content, 'html.parser')
        muster_div = soup6.find('div', {'id': 'RepPr1'})
        if not muster_div: return None
        muster_table = muster_div.find('table')
        if not muster_table: return None
        
        header_row = muster_table.find('tr')
        if not header_row: return None
        header_cols = [th.get_text(strip=True).replace('\u00a0', ' ').strip().lower() for th in header_row.find_all(['th', 'td'])]
        workcode_idx = find_col_idx(header_cols, 'work code')
        if workcode_idx is None: return None

        unique_work_codes = set()
        for row in muster_table.find_all('tr')[1:]:
            cols = row.find_all('td')
            if len(cols) > workcode_idx:
                cell_text = cols[workcode_idx].get_text(strip=True)
                match = re.search(r'(\d+/[A-Z]+/\S+)', cell_text)
                if match:
                    unique_work_codes.add(match.group(1).strip())
        
        return sorted(list(unique_work_codes))

    except requests.RequestException as e:
        print(f"Error getting work codes: {e}")
        return None

def get_attendance_reports(attendance_date, panchayath_name, choice, workcodes=None, progress_callback=None):
    """
    The main logic function to fetch and process attendance data, returning Excel files as BytesIO objects.
    """
    panchayath_name = panchayath_name.upper()
    session = requests.Session()
    
    # This part is duplicated from get_panchayath_and_work_codes.
    # In a real app, you'd likely pass the session and intermediate URLs around
    # to avoid re-fetching, but for simplicity here, we repeat the navigation.
    resp = session.get(BASE_URL, headers=HEADERS)
    soup = BeautifulSoup(resp.content, 'html.parser')
    viewstate_tag = soup.find('input', {'id': '__VIEWSTATE'})
    eventvalidation_tag = soup.find('input', {'id': '__EVENTVALIDATION'})
    viewstategen_tag = soup.find('input', {'id': '__VIEWSTATEGENERATOR'})
    data = {
        '__VIEWSTATE': viewstate_tag['value'],
        '__VIEWSTATEGENERATOR': viewstategen_tag['value'],
        '__EVENTVALIDATION': eventvalidation_tag['value'],
        'ctl00$ContentPlaceHolder1$ddlstate': STATE_VALUE,
        'ctl00$ContentPlaceHolder1$ddl_attendance': attendance_date,
        'ctl00$ContentPlaceHolder1$btn_showreport': 'Show Attendance',
    }
    headers_post = HEADERS.copy(); headers_post['Referer'] = BASE_URL
    resp2 = session.post(BASE_URL, data=data, headers=headers_post)
    soup2 = BeautifulSoup(resp2.content, 'html.parser')
    state_table = get_table_by_id_or_div(soup2)
    karnataka_link = get_link_from_table(state_table, 1, 'KARNATAKA')
    karnataka_url = urljoin(BASE_URL, karnataka_link)
    resp3 = session.get(karnataka_url, headers=HEADERS)
    soup3 = BeautifulSoup(resp3.content, 'html.parser')
    dist_table = get_table_by_id_or_div(soup3)
    ballari_link = get_link_from_table(dist_table, 1, DISTRICT_NAME)
    ballari_url = urljoin(karnataka_url, ballari_link)
    resp4 = session.get(ballari_url, headers=HEADERS)
    soup4 = BeautifulSoup(resp4.content, 'html.parser')
    block_table = get_table_by_id_or_div(soup4)
    siruguppa_link = get_link_from_table(block_table, 1, BLOCK_NAME)
    siruguppa_url = urljoin(ballari_url, siruguppa_link)
    resp5 = session.get(siruguppa_url, headers=HEADERS)
    soup5 = BeautifulSoup(resp5.content, 'html.parser')
    panch_div = soup5.find('div', {'id': 'RepPr1'})
    panch_table = panch_div.find('table')
    panchayath_link = get_panchayath_link(panch_table, panchayath_name)
    if not panchayath_link:
        print("No NMR generated by the Panchayath")
        return None
    panchayath_url = urljoin(siruguppa_url, panchayath_link)
    resp6 = session.get(panchayath_url, headers=HEADERS)
    soup6 = BeautifulSoup(resp6.content, 'html.parser')
    muster_div = soup6.find('div', {'id': 'RepPr1'})
    muster_table = muster_div.find('table')
    header_row = muster_table.find('tr')
    header_cols = [th.get_text(strip=True).replace('\u00a0', ' ').strip().lower() for th in header_row.find_all(['th', 'td'])]
    workcode_idx = find_col_idx(header_cols, 'work code')
    muster_no_idx = find_col_idx(header_cols, 'mustroll no')
    if workcode_idx is None or muster_no_idx is None:
        return None

    rows_to_save = get_muster_roll_rows(muster_table, choice, workcodes, workcode_idx, muster_no_idx)
    if not rows_to_save:
        return None

    # Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 40
    row_cursor = 1
    ws.cell(row=row_cursor, column=1, value="District:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=2, value=DISTRICT_LABEL)
    ws.cell(row=row_cursor, column=3, value="Taluk/Block:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=4, value=TALUK_NAME)
    row_cursor += 1
    ws.cell(row=row_cursor, column=1, value="Panchayath:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=2, value=panchayath_name)
    row_cursor += 2 # More space
    attendance_header_written = False

    # Image-only Excel setup
    img_wb = openpyxl.Workbook()
    img_ws = img_wb.active
    img_ws.column_dimensions['B'].width = 40
    img_ws.column_dimensions['C'].width = 30
    img_ws.column_dimensions['D'].width = 40
    img_row_cursor = 1
    img_bytes_refs = []
    img_ws.cell(row=img_row_cursor, column=1, value="District:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=2, value=DISTRICT_LABEL)
    img_ws.cell(row=img_row_cursor, column=3, value="Taluk/Block:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=4, value=TALUK_NAME)
    img_row_cursor += 1
    img_ws.cell(row=img_row_cursor, column=1, value="Panchayath:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=2, value=panchayath_name)
    img_row_cursor += 2 # More space
    header_row_idx = img_row_cursor
    img_ws.cell(row=header_row_idx, column=1, value='Muster Roll No').font = Font(bold=True)
    img_ws.cell(row=header_row_idx, column=2, value='First Photo').font = Font(bold=True)
    img_ws.cell(row=header_row_idx, column=4, value='Second Photo').font = Font(bold=True)
    img_row_cursor += 1

    # verification_format Excel setup
    rfo_wb = openpyxl.Workbook()
    rfo_ws = rfo_wb.active
    rfo_headers = ['Sl No', 'Officer Designation', 'Verification Day', 'Target', 'No of Photos Actual Checking Done', 'Name of GP', 'Work Code', 'MR No.', 'Accepted Photo (No.)', 'Rejected Photo (No.)', 'Taken By', 'Reason of Rejection', 'Action Taken']
    rfo_ws.append(rfo_headers)

    # Main loop: cache attendance data for each muster roll
    muster_data_cache = {}
    if progress_callback:
        progress_callback(0, "Fetching muster roll data from the web...")

    with ThreadPoolExecutor(max_workers=8) as executor:
        muster_urls = [urljoin(panchayath_url, href) for _, href in rows_to_save]
        results = list(executor.map(fetch_muster_data, muster_urls))
    
    total_photos = 0
    for _, photo_urls, _, _, _ in results:
        if photo_urls:
            total_photos += len(photo_urls)

    total_muster_rolls = len(rows_to_save)
    for i, (muster_url, (attendance_data, photo_urls, work_name, header_cells, taken_by)) in enumerate(zip(muster_urls, results)):
        muster_data_cache[muster_url] = (attendance_data, photo_urls, work_name, header_cells, taken_by)
        img_bytes_list = [download_photo(url) for url in photo_urls] if photo_urls else []
        muster_roll_no = rows_to_save[i][0][muster_no_idx].get_text(strip=True)
        work_code_text = rows_to_save[i][0][workcode_idx].get_text(strip=True)
        
        if progress_callback:
            progress = (i + 1) / total_muster_rolls
            message = f"Parsing Muster Roll: {muster_roll_no} ({i + 1}/{total_muster_rolls})"
            progress_callback(progress, message)

        # Attendance Excel
        if not attendance_header_written and header_cells:
            ws.cell(row=row_cursor, column=1, value="Muster Roll No").font = Font(bold=True)
            for col_idx, header in enumerate(header_cells, 2):
                ws.cell(row=row_cursor, column=col_idx, value=header).font = Font(bold=True)
            ws.cell(row=row_cursor, column=len(header_cells) + 2, value="Photo Taken By").font = Font(bold=True)
            ws.cell(row=row_cursor, column=8, value="First Photo").font = Font(bold=True)
            ws.cell(row=row_cursor, column=10, value="Second Photo").font = Font(bold=True)
            row_cursor += 1
            attendance_header_written = True
        if attendance_data:
            for att_row in attendance_data:
                ws.cell(row=row_cursor, column=1, value=muster_roll_no)
                for col_idx, val in enumerate(att_row, 2):
                    cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                    if col_idx == 4:
                        worker_name = str(val)
                        if '(F)' in worker_name.upper(): cell.fill = FEMALE_FILL
                        elif '(M)' in worker_name.upper(): cell.fill = MALE_FILL
                    if col_idx == 6:
                        attendance_status = str(val)
                        if 'P' in attendance_status.upper(): cell.fill = PRESENT_FILL
                        elif 'A' in attendance_status.upper(): cell.fill = ABSENT_FILL
                ws.cell(row=row_cursor, column=len(att_row) + 2, value=taken_by)
                row_cursor += 1
        
        img_row_start = (row_cursor - len(attendance_data)) if attendance_data else row_cursor
        
        if len(img_bytes_list) > 0 and img_bytes_list[0]:
            img_bytes_list[0].seek(0)
            img = XLImage(img_bytes_list[0])
            ws.add_image(img, f"H{img_row_start}")
        
        if len(img_bytes_list) > 1 and img_bytes_list[1]:
            img_bytes_list[1].seek(0)
            img2 = XLImage(img_bytes_list[1])
            ws.add_image(img2, f"J{img_row_start}")
        elif len(img_bytes_list) == 1:
            ws.cell(row=img_row_start, column=10, value="Only one photo captured")

        image_height_in_rows = 20 
        row_cursor += image_height_in_rows if any(img_bytes_list) else 2
        row_cursor += 5

        # Image-only Excel
        start_img_row = img_row_cursor
        img_ws.cell(row=img_row_cursor, column=1, value=muster_roll_no).font = Font(bold=True, size=18)
        
        if len(img_bytes_list) > 0 and img_bytes_list[0]:
            img_bytes_list[0].seek(0)
            img_bytes_for_imgwb = io.BytesIO(img_bytes_list[0].getbuffer())
            img2 = XLImage(img_bytes_for_imgwb)
            img_ws.add_image(img2, f"B{img_row_cursor}")
            img_bytes_refs.append(img_bytes_for_imgwb)

        if len(img_bytes_list) > 1 and img_bytes_list[1]:
            img_bytes_list[1].seek(0)
            img_bytes_for_imgwb2 = io.BytesIO(img_bytes_list[1].getbuffer())
            img3 = XLImage(img_bytes_for_imgwb2)
            img_ws.add_image(img3, f"D{img_row_cursor}")
            img_bytes_refs.append(img_bytes_for_imgwb2)
        elif len(img_bytes_list) == 1:
            img_ws.cell(row=img_row_cursor, column=4, value="Only one photo captured")

        end_img_row = img_row_cursor + image_height_in_rows - 1
        img_ws.merge_cells(start_row=start_img_row, start_column=1, end_row=end_img_row, end_column=1)
        img_ws.cell(row=start_img_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
        img_row_cursor += image_height_in_rows
        img_row_cursor += 5

        # verification_format Excel
        rfo_ws.append([
            i + 1, '', attendance_date, total_photos, total_photos, panchayath_name,
            work_code_text, muster_roll_no, len(img_bytes_list), 0, taken_by, '', '',
        ])

    wb_bytes, img_wb_bytes, rfo_wb_bytes = save_attendance_excel(wb, img_wb, rfo_wb)
    raw_wb_bytes = save_raw_excel(rows_to_save, panchayath_name, attendance_date, muster_no_idx, workcode_idx, panchayath_url, muster_data_cache)

    sanitized_panchayath = panchayath_name.lower().replace('.', '').replace(' ', '_')
    sanitized_date = attendance_date.replace('/', '_')

    return {
        f"nmr_{sanitized_panchayath}_{sanitized_date}.xlsx": wb_bytes,
        f"nmr_images_{sanitized_panchayath}_{sanitized_date}.xlsx": img_wb_bytes,
        f"verification_format_{sanitized_panchayath}_{sanitized_date}.xlsx": rfo_wb_bytes,
        f"nmr_raw_{sanitized_panchayath}_{sanitized_date}.xlsx": raw_wb_bytes,
    }

if __name__ == "__main__":
    # This part remains for potential CLI usage, but the Streamlit app will use the functions directly.
    pass

